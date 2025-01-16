import openpyxl
from openpyxl import Workbook
from datetime import datetime
import schedule
import time
from twilio.rest import Client

# Twilio configuration (Update with your actual credentials)
account_sid = 'ACafd8209553b32cfe84d5a168951a30d8'
auth_token = 'a8efc897981f82ca3ae7b81122a47820'
twilio_number = '+12674818456'

# Set to keep track of alerted phone numbers
alerted_phone_numbers = set()

# Function to generate XLS file
def generate_xls():
    # Specific data for the provided students
    data = [
        {"name": "Mohammed Shoaib", "usn": "60", "phone": "+919353849223", "checkout_time": "10:30 PM"},
        {"name": "Safwan", "usn": "57", "phone": "+919663498691", "checkout_time": "6:00 PM"},
        {"name": "Mubarak", "usn": "63", "phone": "+916360743449", "checkout_time": "10:45 PM"},
        {"name": "Mohammed Sami A ", "usn": "59", "phone": "+918147900275", "checkout_time": "11:45 AM"}
        # Add more students here if needed
    ]

    # Create a new Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # Create the header row
    headers = ["Student Name", "USN", "Phone Number", "Checkout Time"]
    ws.append(headers)

    # Add data rows
    for student in data:
        ws.append([student["name"], student["usn"], student["phone"], student["checkout_time"]])

    # Save the workbook
    file_name = "hostel_attendance.xlsx"
    wb.save(file_name)
    print(f"Excel file '{file_name}' created successfully.")

# Function to send SMS alert
def send_sms_alert(to_phone_number, student_name):
    client = Client(account_sid, auth_token)
    message = client.messages.create(
        body=f"Alert: {student_name} has checked out after 10:00 PM. Please check the hostel attendance system.",
        from_=twilio_number,
        to=to_phone_number
    )
    print(f"SMS Alert sent to {to_phone_number}. Message SID: {message.sid}")
    alerted_phone_numbers.add(to_phone_number)

# Function to read XLS file and check checkout times
def check_checkout_times():
    file_name = "hostel_attendance.xlsx"
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        student_name, usn, phone_number, checkout_time_str = row
        checkout_time = datetime.strptime(checkout_time_str, "%I:%M %p")
        
        # If the checkout time is after 10:00 PM and the student hasn't been alerted yet
        if checkout_time.strftime("%I:%M %p") >= "10:00 PM" and phone_number not in alerted_phone_numbers:
            send_sms_alert(phone_number, student_name)
            break  # Break after sending a single message

# Generate the XLS file initially
generate_xls()

# Schedule the job every minute to check the time and send alerts if needed
schedule.every(1).minutes.do(check_checkout_times)

print("Scheduler started. Press Ctrl+C to stop.")

# Run the scheduler
try:
    while True:
        schedule.run_pending()
        time.sleep(1)
except KeyboardInterrupt:
    print("Scheduler stopped.")